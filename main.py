from openpyxl import Workbook, load_workbook
import os
import pandas as pd

try:
    os.remove("expense.csv")
except:
    pass
try:
    os.remove("expenses.xlsx")
except:
    pass

wb=load_workbook("expense.xlsx")
ws=wb.active

index=[]
items=[] #list of items w/o editing format
date=[]
quantity=1


for col in ws["A"]:
    index.append(col.value)
for col in ws["D"]:
    date.append(col.value)

for col in ws["E"]:
    items.append(col.value)
    #print(col.value)
    #print(col.row)
    #print(col.coordinate)
    #print(type(col.value))
#print(items)
#print()
itemsDict={}
nameDict={}
weightDict={}
quantityDict={}
tempDict={}



index=int(0)


for item in items:
    #print("\n"+item)
    #print("\n")
    newItem=item
    weight="n.a."
    if "(" in item:
        startPos=item.rfind("(")
        endPos=item.rfind(")")
        weight=item[startPos+1:endPos]
        #print("weight:",weight)
        newItem=item.replace("("+weight+")","")
        weightDict[index]=weight                #dictionary
    else:
        weightDict[index]=weight                #dictionary
    if "x" in item:
        try:
            int(item[item.rfind("x")+1])
            startPos=item.rfind("x")
            quantity=int(item[startPos+1:])
            newItem=newItem.replace("x"+str(quantity),"")
            quantityDict[index]=quantity        #dictionary
        except:
            quantity=1
            quantityDict[index]=quantity        #dictionary
        finally:
            #print("quantity:",quantity)
            pass
    else:
        quantity=1
        quantityDict[index]=quantity        #dictionary
        #print("quantity:",quantity)
    #itemsDict[index]=newItem,quantity,weight
    nameDict[index]=newItem        #dictionary
    
    index+=1
    #print(nameDict)
    #print(quantityDict)
    #print(weightDict)
    #print("\n")
    #itemsDict[index]=nameDict,quantityDict,weightDict
    #print(newItem,"\n"+str(quantity),"\n"+weight,"\n")        
    #print(itemsDict)
    #print("\n")
    #print("\n")


itemsDict["Name"]=nameDict
itemsDict["Quantity"]=quantityDict
itemsDict["Weight"]=weightDict

print(itemsDict)











#for x in itemsDict:
#    print(x.value)
#print("done creating dictionary\n")


data=pd.DataFrame(itemsDict)
#print(data)
#df.to_csv(r"expense.csv",index=False)
data.to_excel("expenses.xlsx")